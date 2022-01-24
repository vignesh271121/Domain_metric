
import xlrd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)
class Domain_chart:

    def report_domain(self):
        exist_path_val = 'C:/Users/vinagend/OneDrive - Cisco/Desktop/test_data/jkl.xlsx'
        new_path_val = 'C:/Users/vinagend/OneDrive - Cisco/Desktop/test_data/bala.xlsx'

        workbook = xlrd.open_workbook(exist_path_val, "rb")
        required_data = []
        sh = workbook.sheet_by_name('kl')
        for rownum in range(sh.nrows):
            row_valaues = sh.row_values(rownum)
            required_data.append(row_valaues[2])
        get_domain_list = list(filter(None, required_data))
        Dict_domain = {}

        for domain_name in get_domain_list:
            if domain_name in Dict_domain.keys():
                values = Dict_domain.get(domain_name)
                values = values + 1
                Dict_domain[domain_name] = values

            else:
                Dict_domain[domain_name] = 1
                print(Dict_domain)

        print(get_domain_list)
        print("=============================================================================")
        print(Dict_domain)

        print("=============================================================================")

        reply_count_list = sorted(Dict_domain.items(), key=lambda x: x[1])
        reply_count_list.reverse()
        result_list_val = dict(reply_count_list)

        key_val = result_list_val.keys()
        values_val = result_list_val.values()
        key_value_list = zip(key_val, values_val)
        print(key_value_list)

        workbook = xlsxwriter.Workbook(new_path_val)
        workbook.add_worksheet('Domain_metric')
        workbook.close()
        Domain_chart.barchat_val(new_path_val, key_value_list)

    def barchat_val(path, get_list_of_key_pair, ):

        wb = load_workbook(path)
        ws = wb['Domain_metric']

        rows = get_list_of_key_pair
        count = 0
        ws.append(('Domain_Name', 'Domain_count'))
        for row in rows:
            count = count+1
            ws.append(row)

        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=count)
        titles = Reference(ws, min_col=1, min_row=2, max_row=count)
        chart = BarChart3D()
        chart.title = "Domain metrics"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)

        ws.add_chart(chart, "E5")
        wb.save(path)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    obj_Domain = Domain_chart()
    obj_Domain.report_domain()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
